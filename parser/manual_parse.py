import sqlite3 as sq
import os
from openpyxl import load_workbook
from config import db_tg, ADMIN_ID, folder

async def parsing_from_bd(bot, chat_id):
    chat_id1, chat_id2 = ADMIN_ID
    try:
        message1 = await bot.send_message(chat_id1, "⏳ Ручной парсинг данных запущен...")
        message2 = await bot.send_message(chat_id2, "⏳ Ручной парсинг данных запущен...")

        with sq.connect(db_tg, timeout=10) as db:
            cur = db.cursor()

            # Получаем список файлов в папке
            files = os.listdir(folder)
            total_files = len(files)

            # Проход по всем файлам в папке
            for file_index, file in enumerate(files, start=1):
                wb = load_workbook(os.path.join(folder, file), data_only=True)
                table_name = os.path.splitext(file)[0]

                # Создание таблицы с новой структурой
                cur.execute(f"""
                    CREATE TABLE IF NOT EXISTS {table_name}(
                        groups TEXT, 
                        subgroup INTEGER NOT NULL DEFAULT 1,
                        day TEXT,
                        time TEXT,
                        ChZn TEXT,
                        para TEXT
                    )""")

                # Проход по всем листам в файле
                for sheetname in wb.sheetnames:
                    sheet = wb[sheetname]
                    # Начальная строка - 7, Начальная колонка - 2
                    row_number = 7
                    columns_number = 3

                    # Получаем строку 7, начиная со 2-й колонки
                    row = sheet[row_number]  # Получаем 7-ю строку

                    # Проход по ячейкам в 7-й строке, начиная со второй колонки
                    for idx, cell in enumerate(row[columns_number - 1:]):  # Индексация с 0, поэтому -1
                        group_name = cell.value
                        if cell.value is not None:
                            print(f'Найдено в листе {sheet.title}, в ячейке {cell.coordinate}')

                            # Обработка групп с подгруппами
                            if idx + 1 < len(row[columns_number - 1:]):
                                next_cell = row[columns_number + idx]
                                if next_cell.value is None:
                                    print(f"В {group_name} есть 2 подгруппа")
                                    await parsing_table(sheet, cell, table_name, group_name, 1, cur)
                                    await parsing_table(sheet, next_cell, table_name, group_name, 2, cur)
                                else:
                                    print(f"В {group_name} Подгрупп не обнаружено")
                                    await parsing_table(sheet, cell, table_name, group_name, 1, cur)

                # Обновление прогресса выполнения
                progress = int((file_index / total_files) * 100)
                await bot.edit_message_text(
                    chat_id=chat_id1,
                    message_id=message1.message_id,
                    text=f"⏳ Ручной парсинг данных запущен: {progress}% завершено"
                )
                await bot.edit_message_text(
                    chat_id=chat_id2,
                    message_id=message2.message_id,
                    text=f"⏳ Ручной парсинг данных запущен: {progress}% завершено"
                )

        # Завершающее сообщение
        await bot.edit_message_text(chat_id=chat_id1, message_id=message1.message_id, text="✅ Обработка всех файлов завершена! (ручной парсинг)")
        await bot.edit_message_text(chat_id=chat_id2, message_id=message2.message_id,
                                    text="✅ Обработка всех файлов завершена! (ручной парсинг)")

    except Exception as e:
        error_message = f"❌ Произошла ошибка: {e}"
        print(error_message)
        await bot.send_message(chat_id1, error_message)
        await bot.send_message(chat_id2, error_message)

async def parsing_table(sheet, cell, table_name, group_name, subgroup, cur):
    days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']
    times = [
        "08:00 - 09:35", "09:45 - 11:20", "11:50 - 13:25",
        "13:35 - 15:10", "15:20 - 16:55", "17:05 - 18:40", "18:50 - 20:25"
    ]

    def find_root_cell(sheet, cell):
        for range_ in sheet.merged_cells.ranges:
            if cell.coordinate in range_:
                min_col, min_row, _, _ = range_.bounds
                return sheet.cell(row=min_row, column=min_col)
        return cell

    def is_merged_4_cells(cell):
        for range_ in sheet.merged_cells.ranges:
            if cell.coordinate in range_:
                min_col, min_row, max_col, max_row = range_.bounds
                width = max_col - min_col + 1
                height = max_row - min_row + 1
                if width * height == 4:
                    return True, min_col, min_row, max_col, max_row, width
        return False, None, None, None, None, None

    def is_merged(cell):
        for range_ in sheet.merged_cells.ranges:
            if cell.coordinate in range_:
                return True
        return False

    row = cell.row + 1
    col = cell.column

    for day in days:
        classes_count = 0
        while classes_count < 7:
            current_cell = sheet.cell(row=row, column=col)
            time_slot = times[classes_count]

            merged, min_col, min_row, max_col, max_row, width = is_merged_4_cells(current_cell)
            if merged:
                root_cell = find_root_cell(sheet, current_cell)
                cell_value = root_cell.value

                if width > 2:
                    cur.execute(f"""
                        INSERT INTO {table_name} (groups, subgroup, ChZn, para, day, time)
                        VALUES (?, ?, ?, ?, ?, ?)""",
                                (group_name, subgroup, 'Числитель', cell_value, day, time_slot))

                    znamenatel_cell = sheet.cell(row=row + 1, column=min_col)
                    znamenatel_value = find_root_cell(sheet, znamenatel_cell).value if is_merged(
                        znamenatel_cell) else znamenatel_cell.value

                    if znamenatel_value:
                        cur.execute(f"""
                            INSERT INTO {table_name} (groups, subgroup, ChZn, para, day, time)
                            VALUES (?, ?, ?, ?, ?, ?)""",
                                    (group_name, subgroup, 'Знаменатель', znamenatel_value, day, time_slot))
                else:
                    cur.execute(f"""
                        INSERT INTO {table_name} (groups, subgroup, ChZn, para, day, time)
                        VALUES (?, ?, ?, ?, ?, ?)""",
                                (group_name, subgroup, 'Числ/Знамен', cell_value, day, time_slot))

                row += 2
            else:
                chislitel_cell = sheet.cell(row=row, column=col)
                znamenatel_cell = sheet.cell(row=row + 1, column=col)

                chislitel_value = find_root_cell(sheet, chislitel_cell).value if is_merged(
                    chislitel_cell) else chislitel_cell.value
                znamenatel_value = find_root_cell(sheet, znamenatel_cell).value if is_merged(
                    znamenatel_cell) else znamenatel_cell.value

                if chislitel_value:
                    cur.execute(f"""
                        INSERT INTO {table_name} (groups, subgroup, ChZn, para, day, time)
                        VALUES (?, ?, ?, ?, ?, ?)""",
                                (group_name, subgroup, 'Числитель', chislitel_value, day, time_slot))
                if znamenatel_value:
                    cur.execute(f"""
                        INSERT INTO {table_name} (groups, subgroup, ChZn, para, day, time)
                        VALUES (?, ?, ?, ?, ?, ?)""",
                                (group_name, subgroup, 'Знаменатель', znamenatel_value, day, time_slot))

                row += 2

            classes_count += 1
        row = cell.row + 1 + ((days.index(day) + 1) * 14)


async def clear_bd():
    if os.path.exists(db_tg):
        os.remove(db_tg)
        return f"Файл базы данных '{db_tg}' успешно удален."
    else:
        return f"Файл базы данных '{db_tg}' не найден."
