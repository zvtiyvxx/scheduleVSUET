import sqlite3 as sq
import os
import hashlib
import shutil
import asyncio
import aiohttp
import aiofiles
from openpyxl import load_workbook
from config import name_table, db_tg, foldercheck, folder, ADMIN_ID
from app.work_schedulesdb import update_unique_schedules


async def download_tables(bot):
    base_url = "https://vsuet.ru/images/student/schedule/"
    os.makedirs(foldercheck, exist_ok=True)

    async with aiohttp.ClientSession() as session:
        for table_name in name_table:
            url = base_url + table_name
            path = os.path.join(foldercheck, table_name)

            os.makedirs(os.path.dirname(path), exist_ok=True)

            try:
                # Проверка типа контента
                async with session.head(url) as resp:
                    content_type = resp.headers.get("Content-Type", "")
                    if "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" not in content_type:
                        # Если это не .xlsx, добавляем "ocz/" в ссылку
                        url = base_url + "ocz/" + table_name

                # Скачиваем файл
                async with session.get(url) as resp:
                    resp.raise_for_status()
                    async with aiofiles.open(path, 'wb') as f:
                        async for chunk in resp.content.iter_chunked(8192):
                            await f.write(chunk)

                print(f"✅ Загружен: {table_name} | Размер: {os.path.getsize(path)} байт")

            except Exception as e:
                print(f"❌ Ошибка при загрузке {table_name}: {e}")

    await asyncio.sleep(1)
    await compare_hashes(bot, ADMIN_ID)

def get_file_hash(path, method='md5'):
    hash_func = hashlib.md5() if method == 'md5' else hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(4096), b''):
            hash_func.update(chunk)
    return hash_func.hexdigest()


async def compare_hashes(bot, id1):
    old_files = {f: get_file_hash(os.path.join(folder, f)) for f in os.listdir(folder) if f.endswith('.xlsx')}
    new_files = {f: get_file_hash(os.path.join(foldercheck, f)) for f in os.listdir(foldercheck) if f.endswith('.xlsx')}
    updated_files = []

    for filename in new_files:
        if filename not in old_files:
            print(f"🆕 Новый файл: {filename}")
            updated_files.append(filename)
            await bot.send_message(ADMIN_ID, "🆕 Новый файл: {filename}")
            return False
        elif new_files[filename] != old_files[filename]:
            print(f"🔄 Обновлён файл: {filename}")
            updated_files.append(filename)
        else:
            print(f"✅ Не изменился: {filename}")

    if updated_files:
        await parsing_from_bd(bot, updated_files)

# Обработка групп и подгрупп, с удалением старых данных и парсингом
async def parsing_from_bd(bot, update_files):
    chat_id1, chat_id2 = ADMIN_ID
    try:
        # Сообщение о начале парсинга
        await bot.send_message(chat_id1, f"📅 Обнаружены измененные таблицы {update_files}, процесс парсинга запущен...")
        await bot.send_message(chat_id2, f"📅 Обнаружены измененные таблицы {update_files}, процесс парсинга запущен...")

        with sq.connect(db_tg, timeout=10) as db:
            cur = db.cursor()

            # Получаем список файлов в папке
            files = os.listdir(foldercheck)

            # Функция для удаления старой таблицы для группы
            def delete_old_schedule(tabl_name):
                cur.execute(f"DROP TABLE IF EXISTS {tabl_name}")

            # Проход по всем файлам в папке
            for file_index, file in enumerate(files, start=1):
                if file in update_files:

                    print(f"Обрабатывается таблица: {file}")
                    wb = load_workbook(os.path.join(foldercheck, file), data_only=True)
                    table_name = os.path.splitext(file)[0]

                    delete_old_schedule(table_name)

                    # Создание таблицы с новой структурой
                    cur.execute(f"""
                        CREATE TABLE IF NOT EXISTS {table_name}(
                            groups TEXT, 
                            subgroup INTEGER NOT NULL DEFAULT 1,
                            day TEXT,
                            time TEXT,
                            ChZn TEXT,
                            para TEXT
                        )""")

                    # Проход по всем листам в файле
                    for sheetname in wb.sheetnames:
                        sheet = wb[sheetname]
                        # Начальная строка - 7, Начальная колонка - 2
                        row_number = 7
                        columns_number = 3

                        # Получаем строку 7, начиная со 2-й колонки
                        row = sheet[row_number]  # Получаем 7-ю строку

                        # Проход по ячейкам в 7-й строке, начиная со второй колонки
                        for idx, cell in enumerate(row[columns_number - 1:]):  # Индексация с 0, поэтому -1
                            group_name = cell.value
                            if group_name is not None:
                                print(f'Найдено в листе {sheet.title}, в ячейке {cell.coordinate}')

                                # Обработка групп с подгруппами
                                if idx + 1 < len(row[columns_number - 1:]):
                                    next_cell = row[columns_number + idx]
                                    if next_cell.value is None:  # Подгруппы 1 и 2
                                        print(f"В {group_name} есть 2 подгруппа")
                                        # Парсим для подгруппы 1 и 2
                                        await parsing_table(sheet, cell, table_name, group_name, 1, cur)
                                        await parsing_table(sheet, next_cell, table_name, group_name, 2, cur)
                                    else:
                                        print(f"В {group_name} подгрупп не обнаружено")
                                        # Парсим только для первой подгруппы
                                        await parsing_table(sheet, cell, table_name, group_name, 1, cur)

        end_check_schedule(update_files)
        await update_unique_schedules()
        # Завершающее сообщение
        await bot.send_message(chat_id=chat_id1,
                                    text="📅 Парсинг завершен, расписание пользователей успешно обновлено!")
        await bot.send_message(chat_id=chat_id2,
                                    text="📅 Парсинг завершен, расписание пользователей успешно обновлено!")


    except Exception as e:
        error_message = f"❌ Произошла ошибка: {e}"
        print(error_message)
        await bot.send_message(chat_id1, error_message)
        await bot.send_message(chat_id2, error_message)

async def parsing_table(sheet, cell, table_name, group_name, subgroup, cur):
    days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']
    times = [
        "08:00 - 09:35", "09:45 - 11:20", "11:50 - 13:25",
        "13:35 - 15:10", "15:20 - 16:55", "17:05 - 18:40", "18:50 - 20:25"
    ]

    def find_root_cell(sheet, cell):
        for range_ in sheet.merged_cells.ranges:
            if cell.coordinate in range_:
                min_col, min_row, _, _ = range_.bounds
                return sheet.cell(row=min_row, column=min_col)
        return cell

    def is_merged_4_cells(cell):
        for range_ in sheet.merged_cells.ranges:
            if cell.coordinate in range_:
                min_col, min_row, max_col, max_row = range_.bounds
                width = max_col - min_col + 1
                height = max_row - min_row + 1
                if width * height == 4:
                    return True, min_col, min_row, max_col, max_row, width
        return False, None, None, None, None, None

    def is_merged(cell):
        for range_ in sheet.merged_cells.ranges:
            if cell.coordinate in range_:
                return True
        return False

    row = cell.row + 1
    col = cell.column

    for day in days:
        classes_count = 0
        while classes_count < 7:
            current_cell = sheet.cell(row=row, column=col)
            time_slot = times[classes_count]

            merged, min_col, min_row, max_col, max_row, width = is_merged_4_cells(current_cell)
            if merged:
                root_cell = find_root_cell(sheet, current_cell)
                cell_value = root_cell.value

                if width > 2:
                    cur.execute(f"""
                        INSERT INTO {table_name} (groups, subgroup, ChZn, para, day, time)
                        VALUES (?, ?, ?, ?, ?, ?)""",
                                (group_name, subgroup, 'Числитель', cell_value, day, time_slot))

                    znamenatel_cell = sheet.cell(row=row + 1, column=min_col)
                    znamenatel_value = find_root_cell(sheet, znamenatel_cell).value if is_merged(
                        znamenatel_cell) else znamenatel_cell.value

                    if znamenatel_value:
                        cur.execute(f"""
                            INSERT INTO {table_name} (groups, subgroup, ChZn, para, day, time)
                            VALUES (?, ?, ?, ?, ?, ?)""",
                                    (group_name, subgroup, 'Знаменатель', znamenatel_value, day, time_slot))
                else:
                    cur.execute(f"""
                        INSERT INTO {table_name} (groups, subgroup, ChZn, para, day, time)
                        VALUES (?, ?, ?, ?, ?, ?)""",
                                (group_name, subgroup, 'Числ/Знамен', cell_value, day, time_slot))

                row += 2
            else:
                chislitel_cell = sheet.cell(row=row, column=col)
                znamenatel_cell = sheet.cell(row=row + 1, column=col)

                chislitel_value = find_root_cell(sheet, chislitel_cell).value if is_merged(
                    chislitel_cell) else chislitel_cell.value
                znamenatel_value = find_root_cell(sheet, znamenatel_cell).value if is_merged(
                    znamenatel_cell) else znamenatel_cell.value

                if chislitel_value:
                    cur.execute(f"""
                        INSERT INTO {table_name} (groups, subgroup, ChZn, para, day, time)
                        VALUES (?, ?, ?, ?, ?, ?)""",
                                (group_name, subgroup, 'Числитель', chislitel_value, day, time_slot))
                if znamenatel_value:
                    cur.execute(f"""
                        INSERT INTO {table_name} (groups, subgroup, ChZn, para, day, time)
                        VALUES (?, ?, ?, ?, ?, ?)""",
                                (group_name, subgroup, 'Знаменатель', znamenatel_value, day, time_slot))

                row += 2

            classes_count += 1
        row = cell.row + 1 + ((days.index(day) + 1) * 14)

def end_check_schedule(update_files):
    try:
        #Перемещаем файлы из временной папки в основную
        for file in update_files:
            #Проверяем, существует ли файл в временной папке
            if os.path.exists(os.path.join(foldercheck, file)):
                shutil.move(os.path.join(foldercheck, file), os.path.join(folder, file))
                print(f"Файл {file} перемещен в основную папку")
            else:
                print(f"Файл {file} не найден в {foldercheck}")

    except Exception as e:
        print(f"Ошибка при обработке файлов: {e}")